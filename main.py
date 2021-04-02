# Manejo de archivos de Excel en python: a partir de  código, podemos crear archivos de excel indicando lo que queremos encontrar en cada celda de la hoja de cálculo
print("Manejo de archivos de Excel:")

# Veremos, en primer lugar, algunas acciones fundamentales para crear nuestro archivo de Excel

# Es necesario importar la librería de Excel para que el código funcione como se espera
import openpyxl

# Crear un archivo de Excel
archivo_exc = openpyxl.Workbook()

# Aquí asingamos una hoja de cálculo en blanco en el archivo creado en el paso anterior
hojacalculo = archivo_exc.active

# Comenzaremos a escribir en nuestra hoja de cálculo de Excel
hojacalculo['A1'] = "Resumen hojas de vida - aspirantes" 

# Guardaremos lo que llevamos con el nombre: primerexcel_py
archivo_exc.save("primerexcel_py.xlsx")

# Abrir archivo
#archivo_exc = openpyxl.load_workbook('primerexcel_py.xlsx')

# Visualizando este archivo, cumple lo que se esperaba, ahora agregaremos más elementos en las celdas
# Agregaremos datos en la columna A: N° aspirante
hojacalculo['A3'] = "N° aspirante"
hojacalculo['A4'] = 1
hojacalculo['A5'] = 2
hojacalculo['A6'] = 3
hojacalculo['A7'] = 4
hojacalculo['A8'] = 5

# Agregaremos datos en la columna B: Nombre
hojacalculo['B3'] = "Nombre"
hojacalculo['B4'] = "Carolina"
hojacalculo['B5'] = "Isabela"
hojacalculo['B6'] = "José David"
hojacalculo['B7'] = "Manuel"
hojacalculo['B8'] = "Susana"

# Agregaremos datos en la columna C: Apellido
hojacalculo['C3'] = "Apellido"
hojacalculo['C4'] = "Gutierrez"
hojacalculo['C5'] = "Orozco"
hojacalculo['C6'] = "Sepúlveda"
hojacalculo['C7'] = "Ortiz"
hojacalculo['C8'] = "Ávila"

# Agregaremos datos en la columna D: Cédula
hojacalculo['D3'] = "Cédula"
hojacalculo['D4'] = 12345
hojacalculo['D5'] = 67891
hojacalculo['D6'] = 23456
hojacalculo['D7'] = 78912
hojacalculo['D8'] = 34567

# Agregaremos datos en la columna E: Edad
hojacalculo['E3'] = "Edad"
hojacalculo['E4'] = 27
hojacalculo['E5'] = 29
hojacalculo['E6'] = 28
hojacalculo['E7'] = 30
hojacalculo['E8'] = 26

# Agregaremos datos en la columna F: Profesión
hojacalculo['F3'] = "Profesión"
hojacalculo['F4'] = "Ingeniera administrativa"
hojacalculo['F5'] = "Administradora de empresas"
hojacalculo['F6'] = "Administrador de empresas"
hojacalculo['F7'] = "Contador público"
hojacalculo['F8'] = "Ingeniera industrial"

# Agregaremos datos en la columna G: Teléfono
hojacalculo['G3'] = "Teléfono"
hojacalculo['G4'] = 5709876
hojacalculo['G5'] = 5754321
hojacalculo['G6'] = 5798765
hojacalculo['G7'] = 5743210
hojacalculo['G8'] = 5712345

# Agregaremos datos en la columna H: Estado civil
hojacalculo['G3'] = "Estado civil"
hojacalculo['G4'] = "Soltera"
hojacalculo['G5'] = "Casada"
hojacalculo['G6'] = "Soltero"
hojacalculo['G7'] = "Casado"
hojacalculo['G8'] = "Soltera"

# Guardaremos lo que llevamos con el nombre: primerexcel_py
archivo_exc.save("primerexcel_py.xlsx")

# Haremos el promedio de edades de los aspirantes
hojacalculo['E10'] = "Promedio edades"
hojacalculo['E11'] = (hojacalculo['E4'].value + hojacalculo['E5'].value + hojacalculo['E6'].value + 
                      hojacalculo['E7'].value + hojacalculo['E8'].value)/5

# Guardaremos este nuevo dato en el documento
archivo_exc.save("primerexcel_py.xlsx")

# Accederemos a algunos datos específicos del archivo
print("\n- Acceder a los valores de celdas específicas")

print()
# Para obtener el dato de una columna y fila derterminada, lo designamos así: hojacalculo.cell(row = 5, column = 2).value
print("Qué dato se encuentra en la fila 5 columna 2?:",hojacalculo.cell(row = 5, column = 2).value)
print("Qué dato se encuentra en la fila 3 columna 4?:",hojacalculo.cell(row = 3, column = 4).value)
print("Qué dato se encuentra en la fila 8 columna 6?:",hojacalculo.cell(row = 8, column = 6).value)
print("Qué dato se encuentra en la fila 4 columna 7?:",hojacalculo.cell(row = 4, column = 7).value)
print("Qué dato se encuentra en la fila 6 columna 5?:",hojacalculo.cell(row = 6, column = 5).value)
print("Qué dato se encuentra en la fila 7 columna 4?:",hojacalculo.cell(row = 7, column = 4).value)
print()
print("El promedio de las edades de los aspirantes es:",hojacalculo['E11'].value)

# Ahora veremos los datos que se encuentran en un cierto rango del archivo
print("\n- Datos desde la celda B4 hasta la D6")

print()
# Creación del for para recorrer las celdas
varias_celdas = hojacalculo['B4':'D6']
for row in varias_celdas:
  for column in row:
    print(column.value,end=" ")

print()
# Ahora veremos los datos que se encuentran en un cierto rango del archivo
print("\n- Datos desde la celda E3 hasta la G7")

print()
# Creación del for para recorrer las celdas
varias_celdas = hojacalculo['E3':'G7']
for row in varias_celdas:
  for column in row:
    print(column.value,end=" ")

print()
# Accederemos a las columnas que hay en el archivo
print("\n- Columnas del archivo")

print()
# Creación del for para recorrer las columnas
todas_col = hojacalculo.columns
for i in todas_col:
  print(i[:])

# Accederemos a las filas que hay en el archivo
print("\n- Filas del archivo")

print()
# Creación del for para recorrer las filas
todas_filas = hojacalculo.rows
for i in todas_filas:
  print(i[:])

# Guardaremos lo que llevamos con el nombre: primerexcel_py
archivo_exc.save("primerexcel_py.xlsx")

# Abrir archivo
archivo_exc = openpyxl.load_workbook('primerexcel_py.xlsx')